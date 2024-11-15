from transformers import pipeline
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

# Δημιουργία του pipeline για text generation με το μοντέλο GPT-2
generator = pipeline('text-generation', model='gpt2')

# Δημιουργία του prompt
prompt = "Can you suggest me a good 32' 4K IPS monitor for programming? Also tell me the price."

# Δημιουργία της απάντησης
response = generator(prompt, max_length=200, num_return_sequences=1)

# Λήψη του κειμένου της απάντησης
result_text = response[0]['generated_text']

# Εκτύπωση της απάντησης στην κονσόλα
print(result_text)

# Ορισμός του ονόματος του αρχείου Excel
excel_filename = 'api_results.xlsx'

# Έλεγχος αν το αρχείο υπάρχει
if os.path.exists(excel_filename):
    # Αν υπάρχει, φόρτωσέ το
    wb = load_workbook(excel_filename)
    sheet = wb.active
else:
    # Αν δεν υπάρχει, δημιούργησε ένα νέο
    wb = Workbook()
    sheet = wb.active
    # Προσθήκη επικεφαλίδων
    sheet['A1'] = 'Timestamp'
    sheet['B1'] = 'Response'

# Λήψη της τρέχουσας ώρας
current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Προσθήκη των νέων δεδομένων στην επόμενη κενή γραμμή
sheet.append([current_time, result_text])

# Αποθήκευση του αρχείου Excel
wb.save(excel_filename)

print(f"Τα αποτελέσματα αποθηκεύτηκαν στο αρχείο {excel_filename}")
